#!/usr/bin/env python

"""ws_telecamere.py: script python che:
    - legge i WS di TEAS 
    - trasferisce i dati sul DB PostgresSQL/PostGIS."""

__author__      = "Roberto Marzocchi"
__copyright__   = "Gter srl"

import os    # standard library
import sys

import requests  # 3rd party packages
import json

import psycopg2
from credenziali import *
conn = psycopg2.connect("dbname={} user={} password={} host={}".format(db, user, pwd, ip))

import openpyxl
from pathlib import Path

from os import listdir
from os.path import isfile, join

# recupero il percorso allo script python 
spath=os.path.dirname(os.path.realpath(__file__)) 
# la sottocartella log deve esserci e possibilmente inseritanel file .gitignore 

# libreria per gestione log
import logging
# inserire riferimento a data e schema 


logging.basicConfig(     
    format='%(asctime)s\t%(levelname)s\t%(message)s',
    filemode='w', # overwrite or append 
    filename='{}/log/letture_file-psa.log'.format(spath),  # nome file (commentandolo viene stampato a schermo
    level=logging.WARNING
    )


from datetime import date, timedelta

oggi=date.today().strftime('%Y-%m-%d')
settimana_prima=(date.today() - timedelta(days=7)).strftime('%Y-%m-%d')

logging.debug(oggi)
logging.debug(settimana_prima)



onlyfiles = [f for f in listdir(path_dati) if isfile(join(path_dati, f))]
#logging.debug(onlyfiles)


curr = conn.cursor()
conn.autocommit = True

i=0
while i < len(onlyfiles):
    logging.debug(onlyfiles[i][0:9])
    if onlyfiles[i][0:9] =='EU LOSE57':
        logging.debug('Ok leggo file {}'.format(onlyfiles[i]))
        xlsx_file = Path('{}/'.format(path_dati), onlyfiles[i])
        wbs_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wbs_obj.active
        logging.debug(sheet)
        for row in sheet.iter_rows(3, sheet.max_row-2):
            #piazzole.append(row[1].value)
            logging.debug(row[0].value)
            # 0 container
            con=row[0].value
            # 1 I/E/T/D
            i_e_t_d=row[1].value
            # 2 length (piedi)
            lun=row[2].value
            # 5 tipo imballaggio
            imb=row[5].value
            # 8 UNNumber
            if row[8].value.strip()=='':
                UNn=0
            else: 
                UNn=row[8].value
            # 9 IMDG
            imdg=row[9].value.strip()
            # 10 giacenza (su questa colonna c'è da fare eventuale update)
            if  row[10].value=='':
                giac=0
            else: 
                giac=row[10].value
            query_insert='''INSERT INTO terra.giacenze_psa
            (id_container, lungh_ft, tipo_imballaggio, un_numb, imdg, giacenza, i_e_t_d)
            VALUES(%s, %s, %s, %s, %s, %s, %s)
            '''
            try:
                curr.execute(query_insert, (con, lun, imb, UNn, imdg, giac, i_e_t_d))
            except Exception as e:
                logging.info(e)
                query_update=''' UPDATE terra.giacenze_psa
                SET lungh_ft=%s, tipo_imballaggio=%s, un_numb=%s, imdg=%s, giacenza=%s, i_e_t_d=%s
                WHERE id_container=%s '''
                try:
                    curr.execute(query_update, (lun, imb, UNn, imdg, giac, i_e_t_d, con))
                except Exception as e:
                    logging.error('File {} - Container {} - problema update'.format(onlyfiles[i], row[0].value))
                    logging.error(e)
    elif onlyfiles[i][0:9] =='EU LOSE45':
        logging.debug('Ok leggo file {}'.format(onlyfiles[i]))
        xlsx_file = Path('{}/'.format(path_dati), onlyfiles[i])
        wbs_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wbs_obj.active
        logging.debug(sheet)
        for row in sheet.iter_rows(3, sheet.max_row-2):
            #piazzole.append(row[1].value)
            logging.debug(row[0].value)
            query_insert=''' INSERT INTO terra.uscite_psa
                (id_container, peso_netto, data_ora_uscita, data_ora_ingresso, tipo_ingresso, tipo_uscita, tipo_contenitore,un_numb, imdg)
                VALUES(%s, %s, %s, %s, %s, %s, %s,%s,%s)'''
            try:
                curr.execute(query_insert, (row[0].value, row[12].value, row[2].value, row[1].value, row[4].value, row[5].value, row[6].value,row[8].value, row[7].value ))
            except Exception as e:
                logging.info(e)
                query_update='''UPDATE terra.uscite_psa
                    SET  peso_netto=%s, data_ora_uscita=%s, 
                    data_ora_ingresso=%s, tipo_ingresso=%s,
                    tipo_uscita=%s, tipo_contenitore=%s, un_numb=%s, imdg=%s 
                    where id_container=%s'''
                try:
                    curr.execute(query_update, (row[12].value, row[2].value, row[1].value, row[4].value, row[5].value, row[6].value, row[8].value, row[7].value, row[0].value))
                except Exception as e:
                    logging.error('Container {} - problema update uscite_psa'.format(row[0].value))   
                    logging.error(e)
    else:   
        logging.debug('Passo oltre con il file {}'.format(onlyfiles[i]))
    i+=1
exit()

